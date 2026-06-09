"""
筛选结果 Excel 导出模块

将 filter_articles.py 的筛选结果导出为格式化的 .xlsx 文件，
方便用户在 Excel/WPS 中查看、排序、筛选。

使用方式：
  1. 被 filter_articles.py 自动调用（筛选完成后）
  2. 独立运行：python view_results.py（查看上次筛选结果）
"""

import os
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import storage


# Excel 输出路径
EXCEL_OUTPUT_FILE = os.path.join(storage.DATA_DIR, "filter_result.xlsx")

# 表头配置：(列名, 对应的字段key, 列宽)
COLUMNS = [
    ("序号",     None,       6),
    ("日期",     "date",    12),
    ("公众号",   "account", 18),
    ("标题",     "title",   40),
    ("链接",     "url",     30),
    ("AI 理由",  "reason",  35),
    ("新增日期", None,      12),
]


def export_to_excel(articles, topic_name=""):
    """将筛选结果导出为格式化的 Excel 文件

    Args:
        articles: 筛选后的文章列表，每项包含 title, account, date, url, reason
        topic_name: 主题名称，用于 sheet 标题

    Returns:
        生成的 Excel 文件绝对路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = topic_name or "筛选结果"

    # ===== 样式定义 =====
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )



    # ===== 写入表头 =====
    for col_idx, (col_name, _, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ===== 写入数据行 =====
    today_str = datetime.date.today().isoformat()

    for row_idx, article in enumerate(articles, start=2):
        seq = row_idx - 1  # 序号从 1 开始

        row_data = [
            seq,
            article.get("date", ""),
            article.get("account", ""),
            article.get("title", ""),
            article.get("url", ""),
            article.get("reason", ""),
            today_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border


        # 序号和日期列居中
        ws.cell(row=row_idx, column=1).alignment = center_alignment
        ws.cell(row=row_idx, column=2).alignment = center_alignment
        ws.cell(row=row_idx, column=7).alignment = center_alignment

        # 链接列设置为可点击超链接
        url = article.get("url", "")
        if url:
            link_cell = ws.cell(row=row_idx, column=5)
            link_cell.hyperlink = url
            link_cell.font = link_font

    # ===== 设置行高 =====
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, len(articles) + 2):
        ws.row_dimensions[row_idx].height = 20

    # ===== 添加自动筛选器 =====
    if articles:
        last_col_letter = get_column_letter(len(COLUMNS))
        last_row = len(articles) + 1
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # ===== 冻结首行 =====
    ws.freeze_panes = "A2"

    # ===== 保存文件 =====
    storage._ensure_data_dir()
    wb.save(EXCEL_OUTPUT_FILE)

    print(f"📊 Excel 已生成：{EXCEL_OUTPUT_FILE}")
    return EXCEL_OUTPUT_FILE


def main():
    """独立运行：读取上次的筛选结果并导出 Excel"""
    if not os.path.exists(storage.FILTER_RESULT_FILE):
        print("⚠️ 未找到筛选结果文件（data/filter_result.json）。")
        print("请先运行 filter_articles.py 进行筛选。")
        return

    with open(storage.FILTER_RESULT_FILE, "r", encoding="utf-8") as f:
        articles = json.load(f)

    if not articles:
        print("⚠️ 筛选结果为空，无文章可导出。")
        return

    print(f"读取到 {len(articles)} 篇筛选结果。")
    export_to_excel(articles)


if __name__ == "__main__":
    main()
